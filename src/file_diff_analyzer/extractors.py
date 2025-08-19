"""
Text extraction module for the File Difference Analyzer library

Handles text extraction from various file formats with fallback mechanisms.
"""

import os
import re
from pathlib import Path
from typing import Optional, Tuple
from .models import FileType


class TextExtractor:
    """Extracts text content from various file formats"""
    
    def __init__(self):
        self._import_optional_dependencies()
    
    def _import_optional_dependencies(self):
        """Import optional dependencies with fallbacks"""
        # PDF extraction
        try:
            import PyPDF2
            self.PyPDF2 = PyPDF2
            self.pdf_supported = True
        except ImportError:
            self.pdf_supported = False
            print("Warning: PyPDF2 not installed. PDF files will not be supported.")
        
        # DOCX extraction
        try:
            import docx
            self.docx = docx
            self.docx_supported = True
        except ImportError:
            self.docx_supported = False
            print("Warning: python-docx not installed. DOCX files will not be supported.")
        
        # Excel extraction
        try:
            import openpyxl
            self.openpyxl = openpyxl
            self.excel_supported = True
        except ImportError:
            self.excel_supported = False
            print("Warning: openpyxl not installed. Excel files will not be supported.")
        
        # CSV extraction
        try:
            import pandas as pd
            self.pandas = pd
            self.csv_supported = True
        except ImportError:
            self.csv_supported = False
            print("Warning: pandas not installed. CSV files will not be supported.")
    
    def extract_text(self, file_path: str) -> Tuple[str, FileType]:
        """Extracts text from a file and determines its type"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        file_path_obj = Path(file_path)
        file_extension = file_path_obj.suffix.lower()
        
        # Determine file type and extract text
        if file_extension == '.pdf':
            if not self.pdf_supported:
                raise ImportError("PDF support not available. Install PyPDF2.")
            return self._extract_pdf_text(file_path), FileType.PDF
        
        elif file_extension == '.docx':
            if not self.docx_supported:
                raise ImportError("DOCX support not available. Install python-docx.")
            return self._extract_docx_text(file_path), FileType.DOCX
        
        elif file_extension in ['.xlsx', '.xls']:
            if not self.excel_supported:
                raise ImportError("Excel support not available. Install openpyxl.")
            return self._extract_excel_text(file_path), FileType.EXCEL
        
        elif file_extension == '.csv':
            if not self.csv_supported:
                # Fallback to basic CSV reading
                return self._extract_csv_text_basic(file_path), FileType.CSV
            return self._extract_csv_text(file_path), FileType.CSV
        
        elif file_extension == '.txt':
            return self._extract_txt_text(file_path), FileType.TXT
        
        else:
            # Try to read as text file
            try:
                return self._extract_txt_text(file_path), FileType.TXT
            except UnicodeDecodeError:
                raise ValueError(f"Unsupported file type: {file_extension}")
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extracts text from PDF file with structure preservation"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = self.PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        # Simple approach: split text into logical lines
                        # Replace common patterns that indicate line breaks
                        
                        # First, normalize the text
                        normalized_text = page_text.replace('\r\n', '\n').replace('\r', '\n')
                        
                        # Split by various patterns that indicate new lines
                        # Pattern 1: Bullet points
                        normalized_text = re.sub(r'([^\n])-\s+', r'\1\n- ', normalized_text)
                        
                        # Pattern 2: Numbered lists
                        normalized_text = re.sub(r'([^\n])(\d+[\.\)]\s+)', r'\1\n\2', normalized_text)
                        
                        # Pattern 3: Headers (capital letters at start)
                        normalized_text = re.sub(r'([a-z])([A-Z][a-zA-Z\s:]{5,})', r'\1\n\2', normalized_text)
                        
                        # Pattern 4: After sentence endings, if next char is uppercase
                        normalized_text = re.sub(r'([.!?])\s+([A-Z])', r'\1\n\2', normalized_text)
                        
                        # Clean up PDF artifacts (file paths, timestamps, etc.)
                        lines = normalized_text.split('\n')
                        cleaned_lines = []
                        for line in lines:
                            line = line.strip()
                            # Skip lines that look like file paths or timestamps
                            if (line.startswith('file:///') or 
                                re.match(r'.*\[\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2}\]', line)):
                                # Extract content after the timestamp if any
                                match = re.search(r'\[\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2}\](.+)$', line)
                                if match:
                                    content = match.group(1).strip()
                                    if content:
                                        cleaned_lines.append(content)
                                continue
                            if line:
                                cleaned_lines.append(line)
                        
                        text += '\n'.join(cleaned_lines) + '\n'
                
                return self._clean_text(text)
        except Exception as e:
            raise RuntimeError(f"Error extracting text from PDF: {e}")
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extracts text from DOCX file"""
        try:
            doc = self.docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return self._clean_text(text)
        except Exception as e:
            raise RuntimeError(f"Error extracting text from DOCX: {e}")
    
    def _extract_excel_text(self, file_path: str) -> str:
        """Extracts text from Excel file"""
        try:
            workbook = self.openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"Sheet: {sheet_name}\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text += row_text + "\n"
                text += "\n"
            return self._clean_text(text)
        except Exception as e:
            raise RuntimeError(f"Error extracting text from Excel: {e}")
    
    def _extract_csv_text(self, file_path: str) -> str:
        """Extracts text from CSV file using pandas"""
        try:
            df = self.pandas.read_csv(file_path)
            text = df.to_string(index=False)
            return self._clean_text(text)
        except Exception as e:
            # Fallback to basic CSV reading
            return self._extract_csv_text_basic(file_path)
    
    def _extract_csv_text_basic(self, file_path: str) -> str:
        """Basic CSV text extraction without pandas"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                return self._clean_text(text)
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    text = file.read()
                    return self._clean_text(text)
            except Exception as e:
                raise RuntimeError(f"Error reading CSV file: {e}")
    
    def _extract_txt_text(self, file_path: str) -> str:
        """Extracts text from text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
                return self._clean_text(text)
        except UnicodeDecodeError:
            # Try with different encoding
            try:
                with open(file_path, 'r', encoding='latin-1') as file:
                    text = file.read()
                    return self._clean_text(text)
            except Exception as e:
                raise RuntimeError(f"Error reading text file: {e}")
    
    def _clean_text(self, text: str) -> str:
        """Cleans and normalizes extracted text"""
        if not text:
            return ""
        
        # Normalize line endings first
        text = text.replace('\r\n', '\n').replace('\r', '\n')
        
        # Remove control characters except newlines and tabs
        text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
        
        # Remove excessive whitespace within lines (but preserve line breaks)
        lines = text.split('\n')
        cleaned_lines = []
        for line in lines:
            # Remove excessive whitespace within each line
            cleaned_line = re.sub(r'[ \t]+', ' ', line).strip()
            cleaned_lines.append(cleaned_line)
        
        text = '\n'.join(cleaned_lines)
        
        return text.strip()
    
    def get_file_info(self, file_path: str) -> Tuple[str, FileType, int]:
        """Gets file information including text content, type, and size"""
        text, file_type = self.extract_text(file_path)
        size_bytes = os.path.getsize(file_path) if os.path.exists(file_path) else 0
        return text, file_type, size_bytes
